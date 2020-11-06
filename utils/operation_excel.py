# 使用 openpyxl
import openpyxl
import sys
# 导入字体、边框、颜色以及对齐方式相关库

from openpyxl.styles import PatternFill


class OperationExcel():
    def __init__(self, file_name=None, sheet_id=None, sheet_name=None):
        if file_name:
            self.file_name = file_name
        else:
            self.file_name = '../data/case.xlsx'
        if sheet_id:
            self.sheet_id = sheet_id
        else:
            self.sheet_id = 0
        if sheet_name:
            self.sheet_name = sheet_name
        else:
            self.sheet_name = "login"

        self.pass_font = PatternFill(fill_type='solid',fgColor="00FF33")
        self.fail_font = PatternFill(fill_type='solid',fgColor="FF0000")
        self.else_font = PatternFill(fill_type='solid',fgColor="FFFF00")

        self.ws = self.__get_sheet_data()

    # 获取sheets的句柄 内容
    def __get_sheet_data(self):
        self.wb = openpyxl.load_workbook(self.file_name)
        # sheets = self.data.get_sheet_by_name(self.sheet_name)
        sheets = self.wb[self.sheet_name]
        return sheets

    # 获取sheet name
    def get_sheet_name(self):

        return self.sheet_name

    # 获取该sheet单元格的总行数
    def get_sheet_rows_num(self):
        return self.ws.max_row

    # 获取该sheet单元格的总列数
    def get_sheet_cols_num(self):
        return self.ws.max_column

    # 获取该sheet单元格的内容
    def get_cell_value(self, row, col):
        try:
            # return self.sheet.cell_value(row,col).value
            return self.ws.cell(row, col).value
        except ValueError as e:
            now_row_col = '. Now row: ' + str(row) + ', col: ' + str(col)
            return str(sys._getframe().f_code.co_name) + "Error : " + str(e) + now_row_col

    # 获取该sheet 指定行的所有内容
    def get_row_value(self, row):
        try:
            columns = self.ws.max_column
            rowdata = []
            for i in range(1, columns + 1):
                cellvalue = self.ws.cell(row=row, column=i).value
                rowdata.append(cellvalue)
            return rowdata
        except ValueError as e:
            now_row_col = '. Now row: ' + str(row)
            return str(sys._getframe().f_code.co_name) + "Error : " + str(e) + now_row_col

    # 获取该sheet 指定列的所有内容
    def get_col_value(self, col):
        try:
            columns = self.ws.max_row
            rowdata = []
            for i in range(1, columns + 1):
                cellvalue = self.ws.cell(row=i, column=col).value
                rowdata.append(cellvalue)
            return rowdata

        except ValueError as e:
            now_row_col = '. Now col: ' + str(col)
            return str(sys._getframe().f_code.co_name) + "Error : " + str(e) + now_row_col

    def write_cell_value(self, row, col, cellvalue,result=None):
        if result:
            if str(result) == "pass":
                self.ws.cell(row,col).fill=self.pass_font
            elif str(result) == "fail":
                self.ws.cell(row, col).fill=self.fail_font
            else:   #yellow
                self.ws.cell(row, col).fill=self.else_font

        try:
            self.ws.cell(row=row, column=col).value = cellvalue
            self.wb.save(self.file_name)
        except PermissionError as  e:
            print("Write error. " + str(e))
        except ValueError as e:
            now_row_col = '. Now row: ' + str(row)
            return str(sys._getframe().f_code.co_name) + "Error : " + str(e) + now_row_col


if __name__ == '__main__':
    import time

    print(time.time())
    ex = OperationExcel(file_name='../data/case.xlsx')
    # ex.write_cell_value(1,1,"ID")
    print(ex.get_cell_value(2, 2))
    print(ex.get_col_value(1))
    # ex.write_cell_value(0,0)
    print(time.time())

# book = xlrd.open_workbook('../data/case.xlsx')  # 得到excel文件的操作对象
# #得到sheet对象
# sheet0 = book.sheet_by_index(0)     #通过sheet索引 得到sheet对象
# sheet_name = book.sheet_names()[0]  #获得指定 索引的sheet名字
# d = book.sheets()[0]
# print(d.ncols)
# print(sheet_name)
#
# sheet1 = book.sheet_by_name('test')        #通过sheet名字来获取，
#
# nrows = sheet0.nrows    #行总数
# ncols = sheet0.ncols
# print(nrows,ncols)
#
# #获得指定的行、列的值，
# row_data = sheet0.row_values(0)
# col_data = sheet0.col_values(0)
# print(row_data,'\n',col_data)
#
# #获取某个表格内容
# cell_value= sheet0.cell_value(0,0)
# print(cell_value)
